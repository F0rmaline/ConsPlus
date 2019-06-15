'''self.tableWidget.setItem(1, 0, QTableWidgetItem(self.GlobalCompanyData[1]))
       self.tableWidget.setItem(3, 0, QTableWidgetItem(self.GlobalCompanyData[2]))
       self.table_3.setItem(0, 0, QTableWidgetItem(self.GlobalResults[0]))  # ОК
       self.table_3.setItem(1, 0, QTableWidgetItem(self.GlobalResults[1]))  # ОК
       self.table_3.setItem(2, 0, QTableWidgetItem(self.GlobalResults[2]))  # ОК
       self.table_3.setItem(3, 0, QTableWidgetItem(self.GlobalResults[3]))  # ОК
       self.table_3.setItem(4, 0, QTableWidgetItem(self.GlobalResults[4]))  # ОК
       self.table_3.setItem(5, 0, QTableWidgetItem(self.GlobalResults[5]))  # ОК
       # self.table.setItem(12, 0, QTableWidgetItem(self.GlobalResults[6])) #Нужно изменить город! КАПЧА
       self.table_3.setItem(6, 0, QTableWidgetItem(self.GlobalResults[7]))  # ОК
       self.table_3.setItem(7, 0, QTableWidgetItem(self.GlobalResults[12]))  # ОК
       self.table_3.setItem(8, 0, QTableWidgetItem(self.GlobalResults[8]))  # ОК
       # self.table.setItem(17, 0, QTableWidgetItem(self.GlobalResults[9])) #OK
       # self.table.setItem(18, 0, QTableWidgetItem(self.GlobalResults[10])) #ОК
       self.table_3.setItem(9, 0, QTableWidgetItem(self.GlobalResults[11]))  # ОК'''

import sys  # sys нужен для передачи argv в QApplication
import os  # Отсюда нам понадобятся методы для отображения содержимого директорий
import zipfile
import glob
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (QTableWidgetItem, QLineEdit, QMessageBox)
import requests
from htmldom import htmldom
import subprocess
import json
import time
from datetime import datetime, date
from bs4 import BeautifulSoup
import re
import io
from pdfminer3.converter import TextConverter
from pdfminer3.pdfinterp import PDFPageInterpreter
from pdfminer3.pdfinterp import PDFResourceManager
from pdfminer3.pdfpage import PDFPage
import openpyxl
from fpdf import FPDF
import fpdf
import threading
path = os.getcwd()
true_path = ""
for i in path:
    if (i != '\\'):
        true_path += i
    else:
        true_path += "/"
print(true_path)
fpdf.SYSTEM_TTFONTS = true_path

import designm1
import designm2
import inff

class InfoWind(QtWidgets.QMainWindow, inff.Ui_MainWindow):

    def __init__(self, parent=None):
        super(InfoWind, self).__init__(parent)
        self.setupUi(self)

class ModalWind(QtWidgets.QMainWindow, designm2.Ui_MainWindow):

    def __init__(self, parent=None):
        super(ModalWind, self).__init__(parent)
        self.setupUi(self)

       # win = ModalWind(self)  # open new window
        # print("AAA")
        #win.show()

        self.tableWidget.setItem(0, 0, QTableWidgetItem(textboxValue11))
        self.tableWidget.setItem(1, 0, QTableWidgetItem(textboxValue12))
        self.tableWidget.setItem(2, 0, QTableWidgetItem(textboxValue13))
        self.tableWidget.setItem(3, 0, QTableWidgetItem(textboxValue14))

        self.table_3.setItem(0, 0, QTableWidgetItem(textboxValue00))  # ОК
        self.table_3.setItem(1, 0, QTableWidgetItem(textboxValue01))  # ОК
        self.table_3.setItem(2, 0, QTableWidgetItem(textboxValue02))  # ОК
        self.table_3.setItem(3, 0, QTableWidgetItem(textboxValue03))  # ОК
        self.table_3.setItem(4, 0, QTableWidgetItem(textboxValue04))  # ОК
        self.table_3.setItem(5, 0, QTableWidgetItem(textboxValue05))  # ОК

        self.table_3.setItem(6, 0, QTableWidgetItem(textboxValue06))
        self.table_3.setItem(7, 0, QTableWidgetItem(textboxValue07))
        self.table_3.setItem(8, 0, QTableWidgetItem(textboxValue08))# ОК
        num1= str(textboxValue09)
        num2 = str(textboxValue010)
        self.lcdNumber.display(num1)
        self.lcdNumber_2.display(num2)
        if trafficLights=="Red":
            self.widget_3.setStyleSheet("image: url(:/newP/r1.png);")
        if trafficLights=="Yellow":
            self.widget_3.setStyleSheet("image: url(:/newP/y1.png);")
        if trafficLights=="Green":
            self.widget_3.setStyleSheet("image: url(:/newP/g1.png);")
        #self.menuCreate_PDF.triggered.connect(self.on_click1)
        self.pushButton.clicked.connect(self.on_click1)
        self.pushButton_2.clicked.connect(self.on_click2)
        #self.table_3.setItem(8, 0, QTableWidgetItem(self.GlobalResults[8]))  # ОК



    def add_image(self, pdf, image_path):
            pdf.image(image_path, x=130, y=8, w=70)

    def coninfo_table(self, pdf, spacing=2):
            data = [['№ п/п', 'Данные контрагента'],
                    ['1', 'Наименование ЮЛ', textboxValue11],
                    ['2', 'ИНН ЮЛ', textboxValue12],
                    ['3', 'ОГРН', textboxValue13],
                    ['4', 'ФИО РУКОВОДИТЕЛЯ ЮЛ', textboxValue14],
                    ['5', 'ИНН РУКОВОДИТЕЛЯ ЮЛ', textboxValueGD]
                    ]

            pdf.set_font('DejaVu', '', 11)

            col_width = pdf.w / 4.5
            row_height = pdf.font_size
            i = 0
            for row in data:
                j = 0
                for item in row:
                    if ((i == 0) and (j == 1)):
                        pdf.cell(col_width * 3.73, row_height * spacing, txt=item, border=1)
                    elif (j == 0):
                        pdf.cell(col_width / 3.5, row_height * spacing, txt=item, border=1)
                    elif (j == 1):
                        pdf.cell(col_width * 1.9, row_height * spacing, txt=item, border=1)
                    else:
                        pdf.cell(col_width * 1.83, row_height * spacing, txt=item, border=1)
                    j += 1
                i += 1
                pdf.ln(row_height * spacing)

    def simple_table(self, pdf, spacing=2):  # 5,6,7,15,16,17,18
            # AnsList = self.GlobalResults
            data = [['№ п/п', 'Критерий', 'Результат'],
                    ['1', 'Количество ЮЛ, в которых руководитель/учредитель ЮЛ является руководителем/учредителем',
                     textboxValue00],
                    ['2', 'В составе ЮЛ дисквалифицированные лица         ', textboxValue01],
                    ['3',
                     'Невозможность участия ФЛ в организации подтверждена в судебном порядке                            ',
                     textboxValue02],
                    ['4', 'Участие в судебных разбирательствах', textboxValue012],
                    ['5', 'Количество сотрудников', textboxValue03],
                    ['6', 'Нахождение в реестре недобросовестных поставщиков', textboxValue04],
                    ['7', 'Налоговая отчётность', textboxValue05],
                    ['8', 'Состояние банкротства', textboxValue06],
                    ['9', 'Регистрация в ЕГРЮЛ в пределах 6 месяцев до момента запроса', textboxValue07],
                    ['10', 'ЮЛ исключено из ЕГРЮЛ', textboxValue08]
                    ]

            pdf.set_font('DejaVu', '', 11)

            col_width = pdf.w / 4.5
            row_height = pdf.font_size
            top = pdf.y
            offset = pdf.x
            i = 0
            for row in data:
                j = 0
                for item in row:
                    if (j == 0):
                        if (i == 1):
                            pdf.cell(col_width / 3.5, 21, txt=item, border=1)
                        elif (i == 3):
                            pdf.cell(col_width / 3.5, 13, txt=item, border=1)
                        elif (i == 6):
                            pdf.cell(col_width / 3.5, 14, txt=item, border=1)
                        elif (i == 9):
                            pdf.cell(col_width / 3.5, 14, txt=item, border=1)
                        else:
                            pdf.cell(col_width / 3.5, row_height * spacing, txt=item, border=1)
                    elif (j == 1):
                        if ((i == 1) or (i == 3) or (i == 6) or (i == 9)):
                            top = pdf.y
                            offset = pdf.x + col_width * 1.9
                            pdf.multi_cell(col_width * 1.9, 7, txt=item, border=1)
                        else:
                            pdf.cell(col_width * 1.9, row_height * spacing, txt=item, border=1)
                    else:
                        if (i == 1):
                            pdf.x = offset
                            pdf.y = top
                            pdf.multi_cell(col_width * 1.83, 21, txt=item, border=1)
                        elif (i == 3):
                            pdf.x = offset
                            pdf.y = top
                            pdf.multi_cell(col_width * 1.83, 13, txt=item, border=1)
                        elif (i == 6):
                            pdf.x = offset
                            pdf.y = top
                            pdf.multi_cell(col_width * 1.83, 14, txt=item, border=1)
                        elif (i == 9):
                            pdf.x = offset
                            pdf.y = top
                            pdf.multi_cell(col_width * 1.83, 14, txt=item, border=1)
                        else:
                            pdf.cell(col_width * 1.83, row_height * spacing, txt=item, border=1)
                    j += 1
                i += 1
                if ((i != 2) and (i != 4) and (i != 7) and (i != 10)):
                    pdf.ln(row_height * spacing)

    def PdfCreate(self, name, inn, ogrn):
            res = "Хорошая компания!"
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('DejaVu', '', 'DejaVuSansCondensed.ttf', uni=True)
            pdf.add_font('DejaVuB', '', 'DejaVuSansCondensed-Bold.ttf', uni=True)
            pdf.add_font('DejaVuS', '', 'DejaVuSerif.ttf', uni=True)
            pdf.set_font('DejaVu', '', 20)

            # Название системы
            pdf.ln(2)
            text = "Система КонтрагентПлюс"
            pdf.cell(200, 10, txt=text, ln=1)

            # Лого Консультант+
            self.add_image(pdf, 'Logo_K.png')

            # Название компании
            pdf.ln(3)
            pdf.set_font('DejaVu', '', 16)
            pdf.ln(3)
            # text = "Название компании"
            # pdf.cell(200, 10, txt=text, ln=1, align="C")

            # Таблица1
            self.coninfo_table(pdf, 2)
            pdf.ln(5)

            # Таблица2
            self.simple_table(pdf)

            # Вывод
            pdf.ln(3)
            pdf.set_font('DejaVu', '', 14)
            res = " "
            res_next = " "
            res_next1 = " "
            if (trafficLights == "Red"):
                res = "Контрагент не соответствует признакам благонадежности,"
                res_next = "          заключать сделки с ним не рекомендуется."
            elif (trafficLights == "Yellow"):
                res = "У контрагента имеются некоторые признаки"
                res_next = "          однодневности/неблагонадежности,"
                res_next1 = "          будьте осмотрительны при заключении сделок с ним."
            else:
                res = "Контрагент соответствует критериям благонадежности. "
            text = "Итог: " + res
            pdf.cell(200, 10, txt=text, ln=1)
            text = res_next
            pdf.cell(200, 10, txt=text, ln=1)
            text = res_next1
            pdf.cell(200, 10, txt=text, ln=1)

            pdf.output("KontrPlus.pdf")

        #self.table_3.setItem(9, 0, QTableWidgetItem(self.GlobalResults[11]))  # ОК
    def on_click1(self):
        name1 = textboxValue11
        inn2 = textboxValue12
        ogrn4 = textboxValue13
        print("on_click1 - pdf")
        self.PdfCreate(name1, inn2, ogrn4)
        path = os.getcwd()
        true_path = ""
        for i in path:
            if (i != '\\'):
                true_path += i
            else:
                true_path += "/"
        true_path += "/KontrPlus.pdf"
        print(true_path)
        os.startfile(true_path)
        #directory1 = QtWidgets.QFileDialog.getExistingDirectory(self)
        #print(directory1)#вылезет диалоговое окно, в котором
                                                                    #можно указать путь. Он по идее сохранится в переменной directory
                    #если пользователь ничего не выбрал, значение будет None

    def ExcelCreate(self, name, inn, ogrn):
        wb = openpyxl.Workbook()
        # wb = openpyxl.load_workbook('1.xlsx') #Открываем тестовый Excel файл
        #AnsList = self.GlobalResults

        array = {'Таблица 1': [
            ['No п/п', '1', '2', '3', '4', '5'],
            ['Данные контрагента', 'Наименование ЮЛ', 'ИНН ЮЛ', 'ОГРН', 'ФИО РУКОВОДИТЕЛЯ ЮЛ', 'ИНН РУКОВОДИТЕЛЯ ЮЛ'],
            ['Результат', textboxValue11, textboxValue12, textboxValue13,
             textboxValue14, textboxValueGD]],
            'Таблица 2': [
                ['No п/п', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10'],
                ['Критерий',
                 'Количество ЮЛ, руководитель/учредитель в ЮЛ которых является руководителем/учредителем',
                 'В составе ЮЛ дисквалифицированные лица',
                 'Невозможность участия ФЛ в организации подтверждена в судебном порядке',
                 'Количество сотрудников',
                 'Нахождение в реестре недобросовестных поставщиков',
                 'Налоговая отчётность',
                 'Состояние банкротства',
                 'Регистрация в ЕГРЮЛ в пределах 6 месяцев до момента запроса',
                 'ЮЛ исключено из ЕГРЮЛ',
                 'Участие в судебных разбирательствах'],
                ['Результат', textboxValue00,textboxValue01, textboxValue02, textboxValue03, textboxValue04, textboxValue05, textboxValue06,
                 textboxValue07, textboxValue08, textboxValue012]
            ]}

        wb.create_sheet('Критерии')  # Создаем лист с названием «Sheet1»
        worksheet = wb['Критерии']  # Делаем его активным
        worksheet.column_dimensions['B'].width = 90
        worksheet.column_dimensions['C'].width = 50
        data = array.get("Таблица 2")
        ind1 = 65
        ind2 = 1
        for lst in data:
            ind2 = 1
            for field in lst:
                ind = chr(ind1) + str(ind2)
                worksheet[ind] = field
                ind2 += 1
            ind1 += 1

        wb.create_sheet('Данные')  # Создаем лист с названием «Sheet1»
        worksheet = wb['Данные']  # Делаем его активным
        # worksheet['']=""
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 60
        data = array.get("Таблица 1")
        ind1 = 65
        ind2 = 1
        for lst in data:
            ind2 = 1
            for field in lst:
                ind = chr(ind1) + str(ind2)
                worksheet[ind] = field
                ind2 += 1
            ind1 += 1

        delete = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(delete)
        wb.save('KontrPlus.xlsx')  # Сохраняем измененный файл

    def on_click2(self):
        name1 = textboxValue11
        inn2 = textboxValue12
        ogrn4 = textboxValue13
        print("on_click2 - excel")
        self.ExcelCreate(name1, inn2, ogrn4)
        path = os.getcwd()
        true_path = ""
        for i in path:
            if (i != '\\'):
                true_path += i
            else:
                true_path += "/"
        true_path += "/KontrPlus.xlsx"
        print(true_path)
        os.startfile(true_path)

class ExampleApp(QtWidgets.QMainWindow, designm1.Ui_MainWindow):

    GlobalCompanyData = []
    GlobalResults = []
    GenDirInn = ""
    TotalPoints = 0
    Light = [False, False, False, False, False, False, False, False, False, False, False, False]
    lightcolor = "Red"
    sem = 0
    textboxValue11=''
    textboxValue12 = ''
    textboxValue13 = ''
    textboxValue00 = ''
    textboxValue01 = ''
    textboxValue02 = ''
    textboxValue03 = ''
    textboxValue04 = ''
    textboxValue05 = ''
    textboxValue06 = ''
    textboxValue07 = ''
    textboxValue08 = ''
    textboxValue09 = ''
    textboxValue010 = ''
    trafficLights = ''
    textboxValueGD=''
    textboxValue012=''

    def AllData(self,name,inn,ogrn):
        if (name):
            query = name
        elif (inn):
            query = inn
        elif (ogrn):
            query = ogrn
        else:
            print("Введите название, ИНН или ОГРН!")
            return []
        url = "https://egrul.nalog.ru"
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        data = {
            'query': query,
            'nameEq': "on"
        }
        res = s.post(url, data)
        ans = res.text
        ans = json.loads(ans)
        token = ans['t']
        url = "https://egrul.nalog.ru/search-result/" + token + "?r=1552544701250&_=1552544701250"
        ans = s.get(url)
        ans = ans.text
        #print(ans)
        ans = json.loads(ans)
        ans = ans['rows']
        #print(type(ans))
        num = 1
        namelist = []
        for an in ans:
            keys = an.keys()
            name = ""
            if ("n" in keys):
                name = an.get("n")
            namelist.append(str(num) + " - " + name)
            num += 1
        for name in namelist:
            print(name)
        if(len(namelist) != 1):
            print("Выберите нужную вам компанию или уточните запрос!")
            my = 0
            while(my <= 0 or my > len(namelist)):
                my = int(input())
                if(my <= 0 or my > len(namelist)):
                    print("Введите правильный номер!")
                else:
                    my = my - 1
                    break
            print(namelist[my])
        else:
            my = 0

        an = ans[my]
        keys = an.keys()
        adr, name, ogrn, inn, date, genName = "","","","","",""
        if ("a" in keys):
            adr = an.get("a")
        if ("n" in keys):
            name = an.get("n")
        if ("o" in keys):
            ogrn = an.get("o")
        if ("i" in keys):
            inn = an.get("i")
        if ("r" in keys):
            date = an.get("r")
        if ("g" in keys):
            gendir = an.get("g")
            lst = gendir.split()
            genName = ""
            Gen = ""
            for i in range(0,len(lst)-1):
                my = lst.pop(0)
                if (":" in my):
                    Gen += " " + my
                    for j in lst:
                        genName += " " + j
                    break
                else:
                    Gen += " " + my
                    
        namelist.append(str(num) + " - " + name)
        num += 1
        myLst = genName.split()
        if (len(myLst) > 3):
            genName = myLst.pop()
        print("Адрес: ", adr,"Название: ", name,"ОГРН: ", ogrn,"ИНН: ", inn,"Дата: ", date, Gen, genName, "\n")
        self.GlobalCompanyData.append(name)
        self.GlobalCompanyData.append(inn)
        self.GlobalCompanyData.append(ogrn)
        self.GlobalCompanyData.append(genName)
        self.GlobalCompanyData.append(adr)
        self.GlobalCompanyData.append(date)
        return

    def extract_text_from_pdf(self, pdf_path):
        resource_manager = PDFResourceManager()
        fake_file_handle = io.StringIO()
        converter = TextConverter(resource_manager, fake_file_handle)
        page_interpreter = PDFPageInterpreter(resource_manager, converter)
     
        with open(pdf_path, 'rb') as fh:
            for page in PDFPage.get_pages(fh, 
                                          caching=True,
                                          check_extractable=True):
                page_interpreter.process_page(page)
     
            text = fake_file_handle.getvalue()
     
        # close open handles
        converter.close()
        fake_file_handle.close()
     
        if text:
            return text
        
    def GenInn(self, name):
        print("Ищу ИНН руководителя!")
        i = 1
        j = 1
        c = 1
        k = 1
        tokens = []
        total = 0
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        while 1:

            data = {
                'query': name,
                'nameEq': "on"
            }
            ans = []
            url = "https://egrul.nalog.ru"
            res = s.post(url, data)
            call = 0
            while res.status_code != 200:
                res = s.post(url, data)
                if (call == 10):
                    print("Превышено число запросов! ИНН")
                    print("Закончил получать ИНН")
                    return("-")
                else:
                    call += 1
            ans = res.text
            ans = json.loads(ans)
            call = 0
            while 't' not in ans: #Ждём пока не получим ответ от сайта
                res = s.post(url, data)
                if (call == 10):
                    print("Превышено число запросов! ИНН")
                    print("Закончил получать ИНН")
                    return("-")
                else:
                    call += 1
                ans = res.text
                ans = json.loads(ans)
                #time.sleep(1)

            token = ans['t']
            url = "https://egrul.nalog.ru/search-result/" + token + "?r=1553770520480&_=1553770520480"
            ans = s.get(url)
            ans = ans.text
            #print(url)
            #print("Это запрос 2",ans)
            ans = json.loads(ans)
            call = 0
            while 'status' in ans:
                if ans['status'] == 'wait':
                    url = "https://egrul.nalog.ru/search-result/" + token + "?r=1553770520480&_=1553770520480"
                    ans = s.get(url)
                    if (call == 10):
                        print("Превышено число запросов! ИНН")
                        print("Закончил получать ИНН")
                        return("-")
                    else:
                        call += 1
                    ans = ans.text
                    #print("Это запрос 2 (Еще раз)", ans)
                    ans = json.loads(ans)
                    #time.sleep(0.3)

            ans = ans['rows']
            for an in ans:
                #print("****************NEW***************")
                #print(an)
                k+=1
                if total == 0:
                    total = an['tot']
                tokens.append(an['t'])
            j+=1
            #time.sleep(0.5)
            if k >= int(total):
                break

        print(tokens)
        token = tokens[0]
        url = "https://egrul.nalog.ru/vyp-request/"+token+"?r=&_=1553240711&_=1553240742811"
        s.get(url)
        url = "https://egrul.nalog.ru/vyp-status/"+token+"?r=&_=1553240711&_=1553240742811"
        res = s.get(url)
        res = json.loads(res.text)
        while res['status'] == 'wait':
            #time.sleep(1)
            res = s.get(url)
            res = json.loads(res.text)

        url = "https://egrul.nalog.ru/vyp-download/"+token+"?r=&_=1553240711&_=1553240742811"
        f=open(r'Site14.pdf',"wb") #открываем файл для записи, в режиме wb
        ufr = requests.get(url) #делаем запрос
        f.write(ufr.content) #записываем содержимое в файл; как видите - content запроса
        f.close()
        print("Загружено!")
        text = self.extract_text_from_pdf('Site14.pdf')
        #print(text)
        text = text.upper()
        res = re.findall(r'СВЕДЕНИЯ О ЛИЦЕ\, ИМЕЮЩЕМ ПРАВО БЕЗ ДОВЕРЕННОСТИ ДЕЙСТВОВАТЬ ОТ ИМЕНИ ЮРИДИЧЕСКОГОЛИЦА', text)
        if (not res):
            print("Нет доверенного лица!")
            return ("Инн не найден!")
        inn = re.findall(r'ИНН\d\d\d\d\d\d\d\d\d\d\d\d', text)
        if (not re.findall(r'ГЕНЕРАЛЬНЫЙ ДИРЕКТОР', text)):
            print("Ген.директор не найден!")
        #print(inn[1])
        res = re.sub(r'ИНН', '', str(inn[0])) #Что нахожу #На что меняю
        return(res)

    def add_image(self, pdf, image_path):
        pdf.image(image_path, x=130, y=8, w=70)
    
    def coninfo_table(self, pdf,spacing=2):
        data = [['№ п/п', 'Данные контрагента'],
                ['1', 'Наименование ЮЛ', self.GlobalCompanyData[0]],
                ['2', 'ИНН ЮЛ', self.GlobalCompanyData[1]],
                ['3', 'ОГРН', self.GlobalCompanyData[2]],
                ['4', 'ФИО РУКОВОДИТЕЛЯ ЮЛ', self.GlobalCompanyData[3]],
                ['5', 'ИНН РУКОВОДИТЕЛЯ ЮЛ', self.GenDirInn]
                ]
     
        pdf.set_font('DejaVu', '', 11)
        
        col_width = pdf.w / 4.5
        row_height = pdf.font_size
        i = 0
        for row in data:
            j = 0
            for item in row:
                if((i == 0) and (j == 1)):
                    pdf.cell(col_width*3.73, row_height*spacing, txt=item, border=1)
                elif (j == 0):
                    pdf.cell(col_width/3.5, row_height*spacing, txt=item, border=1)
                elif (j == 1):
                    pdf.cell(col_width*1.9, row_height*spacing, txt=item, border=1)
                else:
                    pdf.cell(col_width*1.83, row_height*spacing, txt=item, border=1)
                j += 1
            i += 1
            pdf.ln(row_height*spacing)
    
    def simple_table(self, pdf, spacing=2): # 5,6,7,15,16,17,18
        AnsList = self.GlobalResults
        data = [['№ п/п', 'Критерий', 'Результат'],
                ['1', 'Количество ЮЛ, в которых руководитель/учредитель ЮЛ является руководителем/учредителем', AnsList[0]], 
                ['2', 'В составе ЮЛ дисквалифицированные лица         ', AnsList[1]], 
                ['3', 'Невозможность участия ФЛ в организации подтверждена в судебном порядке                            ', AnsList[2]], 
                ['4', 'Участие в судебных разбирательствах', AnsList[12]],
                ['5', 'Количество сотрудников', AnsList[3]],
                ['6', 'Нахождение в реестре недобросовестных поставщиков', AnsList[4]],
                ['7', 'Налоговая отчётность', AnsList[5]],
                ['8', 'Состояние банкротства', AnsList[7]],
                ['9', 'Регистрация в ЕГРЮЛ в пределах 6 месяцев до момента запроса', AnsList[8]],
                ['10', 'ЮЛ исключено из ЕГРЮЛ', AnsList[11]]
                ]
     
        pdf.set_font('DejaVu', '', 11)
     
        col_width = pdf.w / 4.5
        row_height = pdf.font_size
        top = pdf.y
        offset = pdf.x
        i = 0
        for row in data:
            j = 0
            for item in row:
                if (j == 0):
                    if (i == 1):
                        pdf.cell(col_width/3.5, 21, txt=item, border=1)
                    elif (i == 3):
                        pdf.cell(col_width/3.5, 13, txt=item, border=1)
                    elif (i == 6):
                        pdf.cell(col_width/3.5, 14, txt=item, border=1)
                    elif (i == 9):
                        pdf.cell(col_width/3.5, 14, txt=item, border=1)
                    else:
                        pdf.cell(col_width/3.5, row_height*spacing, txt=item, border=1)
                elif (j == 1):
                    if ((i == 1) or (i == 3) or (i == 6) or (i == 9)):
                        top = pdf.y
                        offset = pdf.x + col_width*1.9
                        pdf.multi_cell(col_width*1.9, 7, txt=item, border=1)
                    else:
                        pdf.cell(col_width*1.9, row_height*spacing, txt=item, border=1)
                else:
                    if (i == 1):
                        pdf.x = offset 
                        pdf.y = top
                        pdf.multi_cell(col_width*1.83, 21, txt=item, border=1)
                    elif (i == 3):
                        pdf.x = offset 
                        pdf.y = top
                        pdf.multi_cell(col_width*1.83, 13, txt=item, border=1)
                    elif (i == 6):
                        pdf.x = offset 
                        pdf.y = top
                        pdf.multi_cell(col_width*1.83, 14, txt=item, border=1)
                    elif (i == 9):
                        pdf.x = offset 
                        pdf.y = top
                        pdf.multi_cell(col_width*1.83, 14, txt=item, border=1)
                    else:
                        pdf.cell(col_width*1.83, row_height*spacing, txt=item, border=1)
                j += 1
            i += 1
            if ((i != 2) and (i != 4) and (i != 7) and (i != 10)):
                pdf.ln(row_height*spacing)
                
    def PdfCreate(self, name, inn, ogrn):     
        res = "Хорошая компания!"  
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font('DejaVu', '', 'DejaVuSansCondensed.ttf', uni=True)
        pdf.add_font('DejaVuB', '', 'DejaVuSansCondensed-Bold.ttf', uni=True)
        pdf.add_font('DejaVuS', '', 'DejaVuSerif.ttf', uni=True)
        pdf.set_font('DejaVu', '', 20)
        
        #Название системы
        pdf.ln(2)
        text = "Система КонтрагентПлюс"
        pdf.cell(200, 10, txt=text, ln=1)
        
        #Лого Консультант+
        self.add_image(pdf, 'Logo_K.png')
        
        #Название компании
        pdf.ln(3)
        pdf.set_font('DejaVu', '', 16)
        pdf.ln(3)
        #text = "Название компании"
        #pdf.cell(200, 10, txt=text, ln=1, align="C")
        
        #Таблица1
        self.coninfo_table(pdf, 2)
        pdf.ln(5)
        
        #Таблица2
        self.simple_table(pdf)
        
        #Вывод
        pdf.ln(3)
        pdf.set_font('DejaVu', '', 14)
        res = " "
        res_next = " "
        res_next1 = " "
        if (self.lightcolor == "Red"):
            res = "Контрагент не соответствует признакам благонадежности,"
            res_next = "          заключать сделки с ним не рекомендуется."
        elif (self.lightcolor == "Yellow"):
            res = "У контрагента имеются некоторые признаки"
            res_next = "          однодневности/неблагонадежности,"
            res_next1 = "          будьте осмотрительны при заключении сделок с ним."
        else:
            res = "Контрагент соответствует критериям благонадежности. "
        text = "Итог: " + res
        pdf.cell(200, 10, txt=text, ln=1)
        text = res_next
        pdf.cell(200, 10, txt=text, ln=1)
        text = res_next1
        pdf.cell(200, 10, txt=text, ln=1)
        
        pdf.output("KontrPlus.pdf")

    def ExcelCreate(self, name, inn, ogrn):        
        wb = openpyxl.Workbook()
        #wb = openpyxl.load_workbook('1.xlsx') #Открываем тестовый Excel файл
        AnsList = self.GlobalResults
        
        array = {'Таблица 1': [
                             ['No п/п','1','2','3','4','5'],
                             ['Данные контрагента','Наименование ЮЛ','ИНН ЮЛ','ОГРН','ФИО РУКОВОДИТЕЛЯ ЮЛ','ИНН РУКОВОДИТЕЛЯ ЮЛ'],
                             ['Результат', self.GlobalCompanyData[0], self.GlobalCompanyData[1], self.GlobalCompanyData[2], self.GlobalCompanyData[3],self.GenDirInn]],
                 'Таблица 2': [
                             ['No п/п','1','2','3','4','5','6','7','8','9','10'],
                             ['Критерий',
                              'Количество ЮЛ, руководитель/учредитель в ЮЛ которых является руководителем/учредителем',
                              'В составе ЮЛ дисквалифицированные лица',
                              'Невозможность участия ФЛ в организации подтверждена в судебном порядке',
                              'Количество сотрудников',
                              'Нахождение в реестре недобросовестных поставщиков',
                              'Налоговая отчётность',
                              'Состояние банкротства',
                              'Регистрация в ЕГРЮЛ в пределах 6 месяцев до момента запроса',
                              'ЮЛ исключено из ЕГРЮЛ',
                              'Участие в судебных разбирательствах'],
                             ['Результат',AnsList[0],AnsList[1],AnsList[2],AnsList[3],AnsList[4],AnsList[5],AnsList[7],AnsList[8],AnsList[11],AnsList[12]]
                            ]}
        
        wb.create_sheet('Критерии') #Создаем лист с названием «Sheet1»
        worksheet = wb['Критерии'] #Делаем его активным
        worksheet.column_dimensions['B'].width = 90
        worksheet.column_dimensions['C'].width = 50
        data = array.get("Таблица 2")
        ind1 = 65
        ind2 = 1
        for lst in data:
            ind2 = 1
            for field in lst:
                ind = chr(ind1) + str(ind2)
                worksheet[ind] = field
                ind2 += 1
            ind1 += 1
        
        wb.create_sheet('Данные') #Создаем лист с названием «Sheet1»
        worksheet = wb['Данные'] #Делаем его активным
        #worksheet['']=""
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 60                  
        data = array.get("Таблица 1")
        ind1 = 65
        ind2 = 1
        for lst in data:
            ind2 = 1
            for field in lst:
                ind = chr(ind1) + str(ind2)
                worksheet[ind] = field
                ind2 += 1
            ind1 += 1
               
        delete = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(delete)  
        wb.save('KontrPlus.xlsx') #Сохраняем измененный файл

    def findCompany(self,name):
        i = 1
        word = name  # Название компании
        word = word.upper()
        for filename in glob.glob('files16/*.xml'):
            f = open(filename, 'r')
            if word in f.read():
                self.Light[10] = True
                self.GlobalResults[8] = "Да"
                self.sem += 1
                print("Закончил 16")
                return("Да")
            i = i + 1
        self.Light[10] = False
        self.GlobalResults[10] = "Нет"
        self.sem += 1
        print("Закончил 16")
        return("Нет")

    def zipWork(self, num):
        fantasy_zip = zipfile.ZipFile(str(num) + '.zip')
        fantasy_zip.extractall('files'+str(num))
        fantasy_zip.close()
        print("Я разархивировал файлы!")

    def PdfWork6(self):
        text = self.extract_text_from_pdf('Site6.pdf')
        res = re.findall(r'Сведения о среднесписочной численностиработников юридического лица запредшествующий календарный год\s*\d+', text)
        my = re.findall(r'\d+',res[0])
        my = my[0]
        if (int(my) == 0):
            self.Light[6] = True
        elif (int(my) <= 3):
            self.Light[6] = True
        else:
            self.Light[6] = False
        self.GlobalResults[3] = my
        self.sem += 1
        print("Закончил 6")
        return(my)

    def findFunc(self):
        flag = False
        word = u'Гиперссылка (URL) на набор'
        with open('Html.txt') as file:
            for line in file:
                if (flag):
                    return (line)
                if word in line:
                    flag = True
            file.close()
            return ("Ссылка не найдена!")

    def solve(self, file_name):
        key = "8e864788faaf19a06fe5432376342ac2"
        rc_in = "http://rucaptcha.com/in.php"
        files = {'file': open(file_name, 'rb')}
        data = {"key": key,
                "json": 1}
        r = requests.post(rc_in, data=data, files=files)
        resp = json.loads(r.text)
        #time.sleep(5)
        id = resp['request']
        rc_res = "http://rucaptcha.com/res.php?key=" + key + "&action=get&json=1&id=" + str(id)
        r = requests.get(rc_res)
        r = json.loads(r.text)
        call = 0
        while r['status'] == 0:
            #time.sleep(3)
            r = requests.get(rc_res)
            # print(r.text)
            r = json.loads(r.text)
            if (call == 40):
                print("Я не смог разобрать капчу!")
                return 0
            else:
                call += 1
            if r['request'] == "ERROR_CAPTCHA_UNSOLVABLE":
                return 0
        return r['request']

    # 631201508698
    # 701700793099


    def get_html1(self, url):
        response = requests.get(url)
        return response.text

    def get_html(self,url, data):
        response = requests.get(url, data)
        return response.text

    def captchaPanel(self, html):
        soup = BeautifulSoup(html, 'lxml')
        name = soup.find('div', id='captchaPanel').find('img')
        # print(name)
        return name

    def captchaField(self, html):
        soup = BeautifulSoup(html, 'lxml')
        name = soup.find('div', class_='form-field captcha-field').find('img')
        # print(name)
        return name

    def img_parse(self, imgUrl):
        # print(type(imgUrl))
        url = imgUrl.get('src')
        # print(url)
        fullurl = "https://service.nalog.ru" + url
        # print(fullurl)
        return fullurl

    def download_capcha(self, url):
        p = requests.get(url)
        out = open("capcha1.gif", "wb")
        out.write(p.content)
        out.close()

    def findHref(self, html):
        soup = BeautifulSoup(html, 'lxml')
        name = soup.find('li', class_='js_kladrs-data').find('a', class_='stext').get('href')
        # print(type(name))
        return name

    def findKladr(self, url):
        html = self.get_html1(url)
        # print(html)
        res = self.findHref(html)
        # print(res)
        resurl = 'https://www.alta.ru' + res
        html = self.get_html1(resurl)
        # print(html)
        soup = BeautifulSoup(html, 'lxml')
        name = soup.find('span', class_='js_kladrs-code')
        return (name.text)

    def parseData8(self, html):
        soup = BeautifulSoup(html, 'lxml')
        name = soup.find('p', class_='noRecords')
        # print(name)
        return name

    def parseData16(self, html):
        soup = BeautifulSoup(html, 'lxml')
        name = soup.find('body')
        f = open("Html.txt", 'w')
        f.write(name.text)
        # print(name)
        return name

    def parseData19(self, html):
        soup = BeautifulSoup(html, 'lxml')
        name = soup.find('div', id='SearchReport').find('div')
        return name


    def parseSite1(self,inn,u): #OK
        print("Сайт 1")
        url1 = 'https://service.nalog.ru/mru.do'
        html = self.get_html1(url1)
        data = self.captchaPanel(html)
        img = self.img_parse(data)
        #print(img)
        ctokenhelp = img[46:]
        ctoken = ctokenhelp[:len(ctokenhelp) - 10]
        #print(ctoken)
        self.download_capcha(img)
        print("Капча скачана! 1")
        capcha = self.solve("capcha1.gif")
        print("Капча разобрана! 1", capcha)

        i = 1
        j = 1
        c = 1
        k = 1
        tokens = []
        total = 0
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        print("Начинаю отправку запросов, 1")
        #print(inn)

        data = {
            'fltype': "2",
            'innfl': inn, #701700793099
            'captcha': capcha,
            'captchaToken': ctoken
        }
        ans = []
        url = "https://service.nalog.ru/mru-proc.do"
        res = s.post(url, data)
        call = 0
        while res.status_code != 200:
            res = s.post(url, data)
            if (call == 10):
                print("Превышено число запросов! 1")
                self.GlobalResults[0] = "1"
                self.sem += 1
                print("Закончил 1")
                return("1")
            else:
                call += 1
        ans = res.text
        #print(ans)
        # print(type(ans))
        ans = json.loads(ans)
        res = ans['rows']
        if (not res):
            self.GlobalResults[0] = "1"
            self.sem += 1
            print("Закончил 1")
            return("1")
        res = res[0]
        res = res.get('КолЮЛ')
        if (int(res) > 10):
            self.Light[5] = True
        else:
            self.Light[5] = False
        self.GlobalResults[0] = res
        self.sem += 1
        print("Закончил 1")
        return res

    def parseSite2_2(self, inn): #OK
        print("Сайт 2_2")
        url1 = 'https://service.nalog.ru/disqualified.do'
        html = self.get_html1(url1)
        data = self.captchaPanel(html)
        img = self.img_parse(data)
        #print(img)
        ctokenhelp = img[46:]
        ctoken = ctokenhelp[:len(ctokenhelp) - 10]
        #print(ctoken)
        self.download_capcha(img)
        print("Капча скачана! 2_2")
        capcha = self.solve("capcha1.gif")
        print("Капча разобрана! 2_2", capcha)

        i = 1
        j = 1
        c = 1
        k = 1
        tokens = []
        total = 0
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        print("Начинаю отправку запросов! 2_2")

        data = {
            'orgInn': inn,  # 7804388121
            'captcha': capcha,
            'captchaToken': ctoken
        }
        ans = []
        url = "https://service.nalog.ru/disqualified-ajax.do"
        res = s.post(url, data)
        call = 0
        while res.status_code != 200:
            res = s.post(url, data)
            if (call == 4):
                print("Превышено число запросов! 2_2")
                self.GlobalResults[1] = "-"
                self.sem += 1
                print("Закончил 2_2")
                return("Превышено число запросов!")
            else:
                call += 1
        ans = res.text
        #print(ans)
        # print(type(ans))
        ans = json.loads(ans)
        res = ans['rows']
        if (not res):
            self.Light[2] = False
            self.GlobalResults[1] = "Нет"
            self.sem += 1
            print("Закончил 2_2")
            return ("Нет")
        else:
            self.Light[2] = True
            self.GlobalResults[1] = "Да"
            self.sem += 1
            print("Закончил 2_2")
            return ("Да")

    def parseSite2_1(self, ogrn, inn): #OK
        print("Сайт 2_1")
        url1 = 'https://service.nalog.ru/disfind.do'
        html = self.get_html1(url1)
        data = self.captchaPanel(html)
        img = self.img_parse(data)
        #print(img)
        ctokenhelp = img[46:]
        ctoken = ctokenhelp[:len(ctokenhelp) - 10]
        #print(ctoken)
        self.download_capcha(img)
        print("Капча скачана! 2_1")
        capcha = self.solve("capcha1.gif")
        print("Капча разобрана! 2_1", capcha)

        i = 1
        j = 1
        c = 1
        k = 1
        tokens = []
        total = 0
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        print("Начинаю отправку запросов! 2_1")

        data = {
            'ogrn': ogrn,  # 1027700070518 #1026301177605
            'captcha': capcha,
            'captchaToken': ctoken
        }
        ans = []
        url = "https://service.nalog.ru/disfind-proc.do"
        res = s.post(url, data)
        call = 0
        while res.status_code != 200:
            res = s.post(url, data)
            if (call == 4):
                print("Превышено число запросов! 2_1")
                return (self.parseSite2_2(inn))
            else:
                call += 1
        ans = res.text
        #print(ans)
        # print(type(ans))
        ans = json.loads(ans)
        res = ans['rows']
        if (not res):
            self.Light[2] = False
            print("Закончил 2_1, начал 2_2")
            return (self.parseSite2_2(inn))
        else:
            self.Light[2] = True
            self.GlobalResults[1] = "Да"
            self.sem += 1
            print("Закончил 2_1")
            return ("Да")

    def parseSite3(self, inn, ogrn): #OK
        print("Сайт 3")
        url1 = 'https://service.nalog.ru/svl.do'
        html = self.get_html1(url1)
        data = self.captchaField(html)
        img = self.img_parse(data)
        #print(img)
        ctokenhelp = img[46:]
        ctoken = ctokenhelp[:len(ctokenhelp) - 10]
        #print(ctoken)
        self.download_capcha(img)
        print("Капча скачана! 3")
        capcha = self.solve("capcha1.gif")
        print("Капча разобрана! 3", capcha)

        i = 1
        j = 1
        c = 1
        k = 1
        tokens = []
        total = 0
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        print("Начинаю отправку запросов! 3")

        data = {
            'isForm': "true",  # 1135476176791 5406766509
            'ogrn': ogrn,
            'inn': inn,
            'captcha': capcha,
            'captchaToken': ctoken
        }
        ans = []
        url = "https://service.nalog.ru/svl.do"
        res = s.post(url, data)
        call = 0
        while res.status_code != 200:
            res = s.post(url, data)
            if (call == 4):
                print("Превышено число запросов! 3")
                self.GlobalResults[2] = "-"
                self.sem += 1
                return("Превышено число запросов!")
            else:
                call += 1
        ans = res.text
        # print(ans)
        # print(type(ans))
        try:
            soup = BeautifulSoup(ans, 'lxml')
            name = soup.find('div', class_='panel').find('div', class_='container').find('div', id='notfound')
            if (name.text == 'Информация по данному юридическому лицу не найдена.'):
                self.Light[3] = False
                self.GlobalResults[2] = "Не подтверждена"
                self.sem += 1
                print("Закончил 3")
                return("Не подтверждена")
            else:
                self.Light[3] = True
                self.GlobalResults[2] = "Не подтверждена"
                self.sem += 1
                print("Закончил 3")
                return ("Подтверждена")
        except:
            self.Light[3] = True
            self.GlobalResults[2] = "Подтверждена"
            self.sem += 1
            print("Закончил 3")
            return ("Подтверждена")

    def parseSite6(self, name, inn, ogrn): #OK
        print("Сайт 6")
        if (name):
            findstr = name
        elif (inn):
            findstr = inn
        elif (ogrn):
            findstr = ogrn
        else:
            print ("Введите название, ИНН или ОГРН!")
            self.GlobalResults[3] = "-"
            self.sem += 1
            print("Закончил 6")
            return("")
        i = 1
        j = 1
        c = 1
        k = 1
        tokens = []
        total = 0
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        #print("Начинаю отправку запросов,6")

        data = {
            'query': findstr
        }
        ans = []
        url = "https://rmsp.nalog.ru/search-proc.json"  # "https://egrul.nalog.ru"
        res = s.post(url, data)
        call = 0
        while res.status_code != 200:
            res = s.post(url, data)
            if (call == 4):
                print("Превышено число запросов! 6")
                self.GlobalResults[3] = "-"
                self.sem += 1
                return("Превышено число запросов!")
            else:
                call += 1
        ans = res.text
        # print(ans)
        # print(ans)
        ans = json.loads(ans)
        #print(ans)
        jsondata = ans['data']
        if (not jsondata):
            print("Информация не найдена! 6")
            self.GlobalResults[3] = "-"
            self.sem += 1
            return("Информация не найдена!")
        truedata = jsondata[0]
        token = truedata.get('token')
        # print(token)
        url = "https://rmsp.nalog.ru/excerpt.pdf?token=" + token
        f = open(r'Site6.pdf', "wb")
        ufr = s.get(url)
        f.write(ufr.content)
        f.close()
        print("Загружено!")
        try:
            return self.PdfWork6()
        except:
            self.Light[6] = False
            self.GlobalResults[3] = '-'
            self.sem += 1
            return ("-")

    def parseSite8(self, name, inn): #OK
        print("Сайт 8")
        if (name):
            findstr = name
        elif (inn):
            findstr = inn
        else:
            print("Введите название или ИНН!")
            print("Закончил 8")
            self.sem += 1
            return("")
        url = "http://zakupki.gov.ru/epz/dishonestsupplier/quicksearch/search.html"
        data = {
            "searchString": findstr
        }

        html = self.get_html(url, data)
        res = self.parseData8(html)
        try:
            r = res.text
            self.Light[7] = False
            self.GlobalResults[4] = "Добросовестный"
            self.sem += 1
            print("Закончил 8")
            return("Добросовестный")
        except:
            self.Light[7] = True
            self.GlobalResults[4] = "Недобросовестный"
            self.sem += 1
            print("Закончил 8")
            return("Недобросовестный")


    def parseSite9(self, inn, u): 
        print("Сайт 9")
        url1 = 'https://service.nalog.ru/zd.do'
        html = self.get_html1(url1)
        data = self.captchaField(html)
        img = self.img_parse(data)
        #print(img)
        ctokenhelp = img[46:]
        ctoken = ctokenhelp[:len(ctokenhelp) - 10]
        #print(ctoken)
        self.download_capcha(img)
        print("Капча скачана! 9")
        capcha = self.solve("capcha1.gif")
        print("Капча разобрана! 9", capcha)

        i = 1
        j = 1
        c = 1
        k = 1
        tokens = []
        total = 0
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        print("Начинаю отправку запросов! 9")

        data = {
            'inn': inn,  # 4312141413 #2634040279
            'captcha': capcha,
            'captchaToken': ctoken
        }
        ans = []
        url = "https://service.nalog.ru/zd-proc.do"
        res = s.post(url, data)
        call = 0
        while res.status_code != 200:
            res = s.post(url, data)
            if (call == 4):
                print("Превышено число запросов!")
                self.GlobalResults[5] = "-"
                self.sem += 1
                print("Закончил 9")
                return("Превышено число запросов!")
            else:
                call += 1
        ans = res.text
        # print(ans)
        ans = json.loads(ans)
        jsondata = ans['text']
        # print(jsondata)
        # print(type(jsondata))
        if ("не имеет превышающую 1000 рублей задолженность" in jsondata):
            self.Light[4] = False
            self.GlobalResults[5] = "Имеется"
            self.sem += 1
            print("Закончил 9")
            return ("Имеется")
        else:
            self.Light[4] = True
            self.GlobalResults[5] = "Отсутствует"
            self.sem += 1
            print("Закончил 9")
            return ("Отсутствует")

    def parseSite10(self, u, v): #OK
        print("Сайт 10")
        url1 = 'https://service.nalog.ru/addrfind.do'
        html = self.get_html1(url1)
        data = self.captchaField(html)
        img = self.img_parse(data)
        #print(img)
        ctokenhelp = img[46:]
        ctoken = ctokenhelp[:len(ctokenhelp) - 10]
        #print(ctoken)
        self.download_capcha(img)
        print("Капча скачана! 10")
        capcha = self.solve("capcha1.gif")
        print("Капча разобрана! 10", capcha)

        region = 'Кировская' + '+Область' + '%2C+'
        city = 'Кирово-Чепецк' + '+Город' + '%2C+'
        street = 'Производственная' + '+Улица'
        my1 = region + city + street
        my2 = region + city
        #print('**********************')
        url1 = 'https://www.alta.ru/fias/search/?s_object_rf=' + my1
        url2 = 'https://www.alta.ru/fias/search/?s_object_rf=' + my2
        #print("Улица ", self.findKladr(url1))
        #print("Город ", self.findKladr(url2))

        i = 1
        j = 1
        c = 1
        k = 1
        tokens = []
        total = 0
        s = requests.Session()
        s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
        dom = htmldom.HtmlDom()
        print("Начинаю отправку запросов! 10")

        data = {
            'regionName': '43 - КИРОВСКАЯ ОБЛ',
            'cityName': 'КИРОВО-ЧЕПЕЦК Г',
            'streetName': 'ПРОИЗВОДСТВЕННАЯ УЛ',
            'region': '43',
            'city': self.findKladr(url2),  # Код КЛАДР
            'street': self.findKladr(url1),
            'house': '6',
            'captcha': capcha,
            'captchaToken': ctoken
        }
        ans = []
        url = "https://service.nalog.ru/addr-find-proc.do"
        res = s.post(url, data)
        call = 0
        while res.status_code != 200:
            res = s.post(url, data)
            if (call == 4):
                print("Превышено число запросов!")
                self.GlobalResults[6] = "-"
                self.sem += 1
                print("Закончил 10")
                return("Превышено число запросов!")
            else:
                call += 1
        ans = res.text
        #print(ans)
        ans = json.loads(ans)
        jsondata = ans['rows']
        #print(jsondata)
        res = jsondata[0]
        res = res.get("REG_COUNT")
        #print(res)
        if (int(res) > 50):
            self.Light[8] = True
        elif (int(res) > 10):
            self.Light[8] = True
        elif (int(res) > 5):
            self.Light[8] = True
        else:
            self.Light[8] = False
        self.GlobalResults[6] = res
        self.sem += 1
        print("Закончил 10")
        return(res)

    def parseSite12(self, name, u): #OK
            print("Сайт 12")
            if (name):
                print("Имя введено!")
            else:
                print("Введите имя!")
                self.GlobalResults[7] = "-"
                self.sem += 1
                print("Закончил 12")
                return("")
            url = 'https://bankrot.fedresurs.ru/DebtorsSearch.aspx?Name='

            i = 1
            j = 1
            c = 1
            k = 1
            tokens = []
            total = 0
            s = requests.Session()
            s.headers['Accept'] = 'application/json, text/javascript, */*; q=0.01'
            dom = htmldom.HtmlDom()
            #print("Начинаю отправку запросов")

            data = {

            }
            ans = []
            url = "https://bankrot.fedresurs.ru/DebtorsSearch.aspx?Name=" + name
            res = s.post(url, data)
            call = 0
            while res.status_code != 200:
                res = s.post(url, data)
                if (call == 4):
                    print("Превышено число запросов!")
                    self.GlobalResults[7] = "-"
                    self.sem += 1
                    print("Закончил 12")
                    return("Превышено число запросов!")
                else:
                    call += 1
            ans = res.text
            # print(ans)
            soup = BeautifulSoup(ans, 'lxml')
            name = soup.find('table', class_='bank')
            if ("По заданным критериям не найдено ни одной записи. Уточните критерии поиска" in name.text):
                self.Light[0] = False
                self.GlobalResults[7] = "Не является банкротом"
                self.sem += 1
                print("Закончил 12")
                return ("Не является банкротом")
            else:
                self.Light[0] = True
                self.GlobalResults[7] = "Банкрот"
                self.sem += 1
                print("Закончил 12")
                return ("Банкрот")
                
    def parseSite13(self,name, u): #OK
        print("Сайт 13")
        s = requests.session()
        s.headers['User-Agent'] = "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36"
        s.headers['Accept'] = "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8"
        r = s.get("http://kad.arbitr.ru")

        url = "https://kad.arbitr.ru/Kad/SearchInstances"
        ts = time.time() - 3 * 366 * 24 * 3600 #За последние 3 года
        dateFrom = datetime.utcfromtimestamp(ts).strftime('%Y-%m-%dT00:00:00')
        s.headers['Content-Type'] = "application/json"
        data = {"Page": 1, "Count": 25, "Courts": [], "DateFrom": dateFrom, "DateTo": "null",
                "Sides": [{"Name": name, "Type": 0, "ExactMatch": "false"}], "Judges": [], "CaseNumbers": [],
                "WithVKSInstances": "false"}

        r = s.post(url, json=data)
        call = 0
        while r.status_code != 200:
            r = s.post(url, json=data)
            if (call == 4):
                print("Превышено число запросов!")
                self.GlobalResults[12] = "-"
                self.sem += 1
                print("Закончил 13")
                return("Превышено число запросов!")
            else:
                call += 1
        r = r.text

        dom = htmldom.HtmlDom().createDom(r)

        if dom.find("input[id=documentsTotalCount]"):
            count = dom.find("input[id=documentsTotalCount]").attr('value')
            ret = {'ist':count}
            self.GlobalResults[13] = ret['ist']
        data = {"Page": 1, "Count": 25, "Courts": [], "DateFrom": dateFrom, "DateTo": "null",
                "Sides": [{"Name": name, "Type": 1, "ExactMatch": "false"}], "Judges": [], "CaseNumbers": [],
                "WithVKSInstances": "false"}

        r = s.post(url, json=data)
        r = r.text
        dom = htmldom.HtmlDom().createDom(r)
        if dom.find("input[id=documentsTotalCount]"):
            count = dom.find("input[id=documentsTotalCount]").attr('value')
            ret['otv'] = count
            self.GlobalResults[14] = ret['otv']
        ret = "Истец: " + str(ret['ist']) + " Ответчик: " + str(ret['otv'])
        self.GlobalResults[12] = ret
        self.sem += 1
        print("Закончил 13")
        return ret

    def parseSite14(self, name, inn, ogrn): #OK
        print("Сайт 14")
        datemy = self.GlobalCompanyData[5]
        try:
            self.GenDirInn = self.GenInn(name)
            print(self.GenDirInn)
        except:
            print("Инн ген.директора не был получен!")
        #print(datemy)
        day = int(datemy[0:2])
        month = int(datemy[3:5])
        year = int(datemy[6:10])
        d = date(year, month, day)
        #print("Дата из pdf: ", d)
        now = datetime.date(datetime.now())
        #print("Сегодня: ", now)
        sub = now - d
        #print(sub)
        sub = str(sub).split()
        res = str(sub[0])
        #print(res)
        if (int(res) <= 183):
            self.Light[9] = True
            self.GlobalResults[8] = "Да"
            self.sem += 1
            print("Закончил 14")
            return ("Да")
        else:
            self.Light[9] = False
            self.GlobalResults[8] = "Нет"
            self.sem += 1
            print("Закончил 14")
            return ("Нет")

    def parseSite16(self, name, u):
        print("Сайт 16")
        if (not name):
            print("Введите имя!")
            self.GlobalResults[9] = "-"
            self.sem += 1
            print("Закончил 16")
            return("")
        url = "https://www.nalog.ru/opendata/7707329152-taxoffence/"

        html = self.get_html1(url)
        res = self.parseData16(html)
        #resUrl = findFunc()
        #resUrl = resUrl[:-1]
        #print(resUrl)
        '''f=open('My.zip',"wb") #открываем файл для записи, в режиме wb
        ufr = requests.get(resUrl) #Можно ручками ссылку вставить
        f.write(ufr.content) #записываем содержимое в файл; как видите - content запроса
        f.close()'''
        # print(res.text)
        #self.zipWork(16)
        return (self.findCompany(name))

    def parseSite17(self,name, u):
        print("Сайт 17")
        if (not name):
            print("Введите имя!")
            self.GlobalResults[10] = "-"
            self.sem += 1
            print("Закончил 17")
            return("")
        ressum = 0
        findflag = False
        findstr = name  # Уральский центр безопасности труда #Слово и дело #Тридос
        findstr = findstr.upper()
        #self.zipWork(17)
        for filename in glob.glob('files17/*.xml'):
            f = open(filename, 'r')
            string = f.read()
            lst = string.split("><")
            # print(lst)
            #print(findstr)
            for i in range(0, len(lst) - 1):
                record = lst.pop(0)
                if findstr in record:
                    #print("Yes")
                    findflag = True
                    while ("Документ" not in record):
                        i += 1
                        record = lst.pop(0)
                        if ("ОбщСумНедоим" in record):
                            #print("Тут!")
                            record = record[record.find("ОбщСумНедоим=") + 14:]
                            record = record[:len(record) - 2]
                            #print(record)
                            ressum += float(record)
                    break
            #print("Итого: ", ressum)
            if (findflag):
                if (ressum <= 300000.0):
                    self.Light[11] = True
                    self.GlobalResults[10] = "Небольшая задолженность"
                    self.sem += 1
                    print("Закончил 17")
                    return("Небольшая задолженность")
                else:
                    self.Light[11] = True
                    self.GlobalResults[10] = "Имеется крупная задолженность"
                    self.sem += 1
                    print("Закончил 17")
                    return("Имеется крупная задолженность")
            else:
                self.Light[11] = False
                #print ("Компания не найдена!")
                #print(self.sem)
        self.GlobalResults[10] = "-"
        self.sem += 1
        print("Закончил 17")
        return ("Компания не найдена!")

    def parseSite19(self,inn, u): #OK
        print("Сайт 19")
        resStr = ""
        if (not inn):
            print("Введите ИНН!")
            self.GlobalResults[11] = "-"
            self.sem += 1
            print("Закончил 19")
            return("")
        url = "https://www.vestnik-gosreg.ru/publ/fz83/"
        data = {
            "query": inn  # 0276115023    7728029110
        }

        html = self.get_html(url, data)
        try:
            res = self.parseData19(html)
            if ("Информации не обнаружено" in res.text):
                resStr = ("Не исключено")
                self.Light[1] = False
                self.GlobalResults[11] = resStr
                self.sem += 1
                print("Закончил 19")
                return (resStr)
            else:
                self.TotalPoints += 3
                resStr = ("Исключено")
                self.Light[1] = True
                self.GlobalResults[11] = resStr
                self.sem += 1
                print("Закончил 19")
                return (resStr)
        except:
            self.GlobalResults[11] = "Не исключено"
            self.sem += 1
            print("Закончил 19")
            return ("Ошибка! Не исключено")
            
    def Analyze(self):
        #print(self.Light[5:])
        count = 0
        for l in self.Light[5:]:
            if l:
                count += 1
        if (self.Light[0] or self.Light[1]):
            return("Red")
        elif (self.Light[2] and self.Light[3] and self.Light[4]):
            return("Red")
        elif (((self.Light[2] and self.Light[3]) or (self.Light[2] and self.Light[4]) or (self.Light[3] and self.Light[4])) and (True in self.Light[5:])):
            return("Red")
        elif (self.Light[2] or self.Light[3] or self.Light[4] or (count >= 3)):
            return("Yellow")
        else:
            return("Green")

    def __init__(self):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна
        #self.pushButton.clicked.connect(self.onChanged)
        self.statusBar.showMessage('(c) Dubrovin M., Korotkov B., Demyanova A.,Grehneva O., Dokuchaeva D., Kalinina A., Consultant Plus, 2019')
        #self.lineEdit.textChanged[str].connect(self.onChanged)
        #self.lineEdit_2.textChanged[str].connect(self.onChanged2)
        #self.lineEdit_5.textChanged[str].connect(self.onChanged5)
        #self.lineEdit_3.textChanged[str].connect(self.onChanged3)
        self.pushButton_2.clicked.connect(self.on_show)
        self.pushButton.clicked.connect(self.on_click)
        #self.pushButtonpdf.clicked.connect(self.on_click1) #!!! сюда вставь название функции
        #self.pushButtonex.clicked.connect(self.on_click2)  # и сюда тоже, да. сейчас эти функции ниже
        #self.pushButton.clicked.connect(self.browse_folder)  # Выполнить функцию browse_folder
                                                            # при нажатии кнопки
        #self.pushButton.clicked.connect(self.on_show)

    def on_show(self):
        win = InfoWind(self)
        # print("AAA")
        win.show()

    def onChanged(self,text):
       # self.textBrowser.setText(text)
        self.table.setItem(0, 0,QTableWidgetItem(text))
    def onChanged2(self,text):
       # self.textBrowser.setText(text)
        self.table.setItem(1, 0, QTableWidgetItem(text))
    def onChanged5(self,text):
       # self.textBrowser.setText(text)
        self.table.setItem(2, 0, QTableWidgetItem(text))
    def onChanged3(self,text):
       # self.textBrowser.setText(text)
        self.table.setItem(3, 0, QTableWidgetItem(text))

    def on_click1(self):
        name1 = self.lineEdit.text()
        inn2 = self.lineEdit_2.text() 
        ogrn4 = self.lineEdit_3.text()
        print("on_click1 - pdf")
        self.PdfCreate(name1, inn2, ogrn4)
        path = os.getcwd()
        true_path = ""
        for i in path:
            if (i != '\\'):
                true_path += i
            else:
                true_path += "/"
        true_path += "/KontrPlus.pdf"
        print(true_path)
        os.startfile(true_path)
        #directory1 = QtWidgets.QFileDialog.getExistingDirectory(self) #вылезет диалоговое окно, в котором
                                                                    #можно указать путь. Он по идее сохранится в переменной directory
                    #если пользователь ничего не выбрал, значение будет None


    def on_click2(self):
        name1 = self.lineEdit.text()
        inn2 = self.lineEdit_2.text()
        ogrn4 = self.lineEdit_3.text()
        print("on_click2 - excel")
        self.ExcelCreate(name1, inn2, ogrn4)
        path = os.getcwd()
        true_path = ""
        for i in path:
            if (i != '\\'):
                true_path += i
            else:
                true_path += "/"
        true_path += "/KontrPlus.xlsx"
        print(true_path)
        os.startfile(true_path)
        #directory2 = QtWidgets.QFileDialog.getExistingDirectory(self) #здесь также, как выше

    def on_click(self):
        self.sem = 0
        global textboxValue11
        global textboxValue12
        global textboxValue13
        global textboxValue14
        global textboxValue00
        global textboxValue01
        global textboxValue02
        global textboxValue03
        global textboxValue04
        global textboxValue05
        global textboxValue06
        global textboxValue07
        global textboxValue08
        global textboxValue09
        global textboxValue010
        global trafficLights
        global textboxValueGD
        global textboxValue012

        self.GlobalCompanyData = []
        self.GlobalResults = [" ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " "," "]
        self.GenDirInn = ""
        self.TotalPoints = 0
        textboxValue1 = self.lineEdit.text()
        textboxValue2 = self.lineEdit_2.text()
        textboxValue3 = self.lineEdit_3.text()

        #textboxValue6 = self.lineEdit_6.text()
        ans1 = ans2 = ans3 = ans4 = ans5 = ans6 = ans7 = ans8 = ans9 = ans10 = ans11 = ans12= ans13 = "-" 
        try:
            self.AllData(textboxValue1, textboxValue2, textboxValue3)
            name = self.GlobalCompanyData[0]
            inn = self.GlobalCompanyData[1]
            ogrn = self.GlobalCompanyData[2]
            #self.table.setItem(0, 0, QTableWidgetItem(self.GlobalCompanyData[0]))
            #self.table.setItem(1, 0, QTableWidgetItem(self.GlobalCompanyData[1]))
            #self.table.setItem(3, 0, QTableWidgetItem(self.GlobalCompanyData[2]))
        except:
            print("Данные не были получены! Попробуйте повторить запрос!")
            return
        try:
            ans14 = self.parseSite14(name, inn, ogrn)
            #self.table.setItem(15, 0, QTableWidgetItem(self.GlobalResults[8])) #ОК
            #self.GlobalResults.append(ans14)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        try: #OK
            t2 = threading.Thread(target=self.parseSite2_1, args=(ogrn,inn))
            t2.start()
            #self.table.setItem(5, 0, QTableWidgetItem(self.GlobalResults[1])) #ОК
            #ans2 = self.parseSite2_1(ogrn,inn)
            #self.GlobalResults.append(ans2)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        try: #OK
            t3 = threading.Thread(target=self.parseSite3, args=(ogrn,inn))
            t3.start()
            #self.table.setItem(6, 0, QTableWidgetItem(self.GlobalResults[2])) #ОК
            #ans3 = self.parseSite3(ogrn, inn)
            #self.GlobalResults.append(ans3)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        try: #OK
            t6 = threading.Thread(target=self.parseSite6, args=(name, inn, ogrn))
            t6.start()
            #self.table.setItem(8, 0, QTableWidgetItem(self.GlobalResults[3])) #ОК
            #ans6 = self.parseSite6(name, inn, ogrn)
            #self.GlobalResults.append(ans6)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        try: #OK
            t8 = threading.Thread(target=self.parseSite8, args=(name, inn))
            t8.start()
            #self.table.setItem(10, 0, QTableWidgetItem(self.GlobalResults[4])) #ОК
            #ans8 = self.parseSite8(name, inn)
            #self.GlobalResults.append(ans8)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        try: #OK
            t9 = threading.Thread(target=self.parseSite9, args=(inn,""))
            t9.start()
            #self.table.setItem(11, 0, QTableWidgetItem(self.GlobalResults[5])) #ОК
            #ans9 = self.parseSite9(inn)
            #self.GlobalResults.append(ans9)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        '''try: #OK
            t10 = threading.Thread(target=self.parseSite10, args=("",""))
            t10.start()
            #self.table.setItem(12, 0, QTableWidgetItem(self.GlobalResults[6])) #Нужно изменить город! КАПЧА
            #ans10 = self.parseSite10()
            #self.GlobalResults.append(ans10)
        except:
            print("Пропускаем сайт")'''
        try: #OK
            t12 = threading.Thread(target=self.parseSite12, args=(name,""))
            t12.start()
            #self.table.setItem(13, 0, QTableWidgetItem(self.GlobalResults[7])) #ОК
            #ans12 = self.parseSite12(name)
            #self.GlobalResults.append(ans12)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        try:
            t1 = threading.Thread(target=self.parseSite1, args=(self.GenDirInn,""))
            t1.start()
            #self.table.setItem(4, 0, QTableWidgetItem(self.GlobalResults[0])) #ОК
            #ans1 = self.parseSite1(self.GenDirInn)
            #self.GlobalResults.insert(0,ans1)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        '''try:
            t16 = threading.Thread(target=self.parseSite16, args=(name,""))
            t16.start()
            self.table.setItem(17, 0, QTableWidgetItem(self.GlobalResults[9])) #OK
            #ans16 = self.parseSite16(name)
            #self.GlobalResults.append(ans16)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        try:
            t17 = threading.Thread(target=self.parseSite17, args=(name,""))
            t17.start()
            self.table.setItem(18, 0, QTableWidgetItem(self.GlobalResults[10])) #ОК
            #ans17 = self.parseSite17(name)
            #self.GlobalResults.append(ans17)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")'''
        try: #OK
            t19 = threading.Thread(target=self.parseSite19, args=(inn,""))
            t19.start()
            #self.table.setItem(19, 0, QTableWidgetItem(self.GlobalResults[11]))
            #ans19 = self.parseSite19(inn) # 0276115023
            #self.GlobalResults.append(ans19)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        try: #OK
            t13 = threading.Thread(target=self.parseSite13, args=(name,""))
            t13.start()
            #self.table.setItem(14, 0, QTableWidgetItem(self.GlobalResults[12])) #ОК
            #ans13 = self.parseSite13(name)
            #ans13 = "Истец: " + str(ans13['ist']) + " Ответчик: " + str(ans13['otv'])
            #self.GlobalResults.append(ans13)
        except:
            self.GlobalResults.append("-")
            print("Пропускаем сайт")
        #while(len(self.GlobalResults) != 4):
            #pass
        #self.GlobalResults.pop(1)
        while(self.sem != 10):
            print(self.sem)
        print(self.GlobalCompanyData)
        print(self.GlobalResults)
        print(self.TotalPoints)
        print(self.Light)
        self.lightcolor = self.Analyze()
        print(self.lightcolor)
        #global textboxValue11
        textboxValue11 = self.GlobalCompanyData[0]
        textboxValue12 = self.GlobalCompanyData[1]
        textboxValue13 = self.GlobalCompanyData[2]
        textboxValue14 = self.GlobalCompanyData[3]
        textboxValue00 = self.GlobalResults[0]
        textboxValue01 = self.GlobalResults[1]
        textboxValue02 = self.GlobalResults[2]
        textboxValue03 = self.GlobalResults[3]
        textboxValue04 = self.GlobalResults[4]
        textboxValue05 = self.GlobalResults[5]
        textboxValue06 = self.GlobalResults[7]
        textboxValue07 = self.GlobalResults[8]
        textboxValue08 = self.GlobalResults[11]
        textboxValue012 = self.GlobalResults[12]
        textboxValue09 = self.GlobalResults[13]
        textboxValue010 = self.GlobalResults[14]
        trafficLights = self.lightcolor
        textboxValueGD = self.GenDirInn
        win = ModalWind(self)
        # print("AAA")
        win.show()

        '''self.table.setItem(4, 0, QTableWidgetItem(self.GlobalResults[0])) #ОК
        self.table.setItem(5, 0, QTableWidgetItem(self.GlobalResults[1])) #ОК
        self.table.setItem(6, 0, QTableWidgetItem(self.GlobalResults[2])) #ОК
        self.table.setItem(8, 0, QTableWidgetItem(self.GlobalResults[3])) #ОК
        self.table.setItem(10, 0, QTableWidgetItem(self.GlobalResults[4])) #ОК
        self.table.setItem(11, 0, QTableWidgetItem(self.GlobalResults[5])) #ОК
        #self.table.setItem(12, 0, QTableWidgetItem(self.GlobalResults[6])) #Нужно изменить город! КАПЧА
        self.table.setItem(13, 0, QTableWidgetItem(self.GlobalResults[7])) #ОК
        self.table.setItem(14, 0, QTableWidgetItem(self.GlobalResults[12])) #ОК
        self.table.setItem(15, 0, QTableWidgetItem(self.GlobalResults[8])) #ОК
        #self.table.setItem(17, 0, QTableWidgetItem(self.GlobalResults[9])) #OK
        #self.table.setItem(18, 0, QTableWidgetItem(self.GlobalResults[10])) #ОК
        self.table.setItem(19, 0, QTableWidgetItem(self.GlobalResults[11])) #ОК
        #self.widget.setStyleSheet("\n" "background-image: url(:/r1.png);")'''

    def browse_folder(self):

        directory = QtWidgets.QFileDialog.getExistingDirectory(self)
        # открыть диалог выбора директории и установить значение переменной
        # равной пути к выбранной директории



def main():
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    app.exec_()  # и запускаем приложение

if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()  # то запускаем функцию main()
